from rest_framework import generics
from rest_framework.permissions import (
    IsAuthenticated,
    SAFE_METHODS,
)
from rest_framework.filters import SearchFilter
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.http import HttpResponse

from accounts.permissions import IsSuperuser

from .models import DiePrice
from .serializers import DiePriceSerializer


# Lists all die prices and handles creating new ones.
# GET  /api/masters/die-prices/  — returns newest die prices first
# POST /api/masters/die-prices/  — creates a new die price entry
class DiePriceListCreateView(
    generics.ListCreateAPIView
):

    serializer_class = DiePriceSerializer
    permission_classes = [IsAuthenticated]

    # Enables searching die prices by SL number or Die No.
    filter_backends = [SearchFilter]

    search_fields = [
        "die_code",
        "name",
    ]

    def get_permissions(self):
        if self.request.method in SAFE_METHODS:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsSuperuser()]

    def get_queryset(self):
        # select_related avoids extra DB queries for created_by
        return DiePrice.objects.select_related(
            "created_by"
        ).order_by("-created_at", "-id")  # Newest first

    def perform_create(self, serializer):
        # Automatically set created_by to the logged-in staff member
        serializer.save(
            created_by=self.request.user
        )


# Retrieve, update, or delete a single die price by ID.
# GET/PATCH/DELETE /api/masters/die-prices/<id>/
class DiePriceDetailView(
    generics.RetrieveUpdateDestroyAPIView
):

    serializer_class = DiePriceSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.request.method in SAFE_METHODS:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsSuperuser()]

    def get_queryset(self):

        return DiePrice.objects.select_related(
            "created_by"
        )


class DiePriceExcelImportView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated, IsSuperuser]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Please select an Excel file."}, status=status.HTTP_400_BAD_REQUEST)
        if not upload.name.lower().endswith((".xlsx", ".xlsm")):
            return Response({"detail": "Only .xlsx or .xlsm files are supported."}, status=status.HTTP_400_BAD_REQUEST)
        if upload.size > 10 * 1024 * 1024:
            return Response({"detail": "Excel file must be smaller than 10 MB."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(upload, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        except Exception:
            return Response({"detail": "Unable to read this Excel file."}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({"detail": "The Excel sheet is empty."}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(value or "").strip().lower().replace(" ", "_") for value in rows[0]]
        # Accept the client-facing die_no column while retaining compatibility
        # with older spreadsheets that used name.
        if "die_no" in headers and "name" not in headers:
            headers[headers.index("die_no")] = "name"
        required = {"name", "rate"}
        missing = sorted(required - set(headers))
        if missing:
            return Response({"detail": f"Missing required column(s): {', '.join(missing)}. Use die_no, rate, and optional die_code/is_active."}, status=status.HTTP_400_BAD_REQUEST)

        column = {name: index for index, name in enumerate(headers)}
        prepared = []
        errors = []
        seen_names = set()
        seen_codes = set()
        existing_names = {
            value.casefold()
            for value in DiePrice.objects.values_list("name", flat=True)
        }
        existing_codes = set(
            DiePrice.objects.values_list("die_code", flat=True)
        )
        next_code_number = (DiePrice.objects.order_by("-id").values_list("id", flat=True).first() or 0) + 1

        for row_number, values in enumerate(rows[1:], start=2):
            if not any(value not in (None, "") for value in values):
                continue
            name = str(values[column["name"]] or "").strip()
            if name.isdigit():
                name = f"DIE - {name}"
            raw_rate = values[column["rate"]]
            code = str(values[column["die_code"]] or "").strip() if "die_code" in column else ""
            active_value = values[column["is_active"]] if "is_active" in column else True
            active_text = str(active_value).strip().lower()
            is_active = active_value is not None and active_text not in {"false", "0", "no", "inactive"}

            if len(name) < 2:
                errors.append(f"Row {row_number}: name is required.")
                continue
            if name.casefold() in seen_names or name.casefold() in existing_names:
                errors.append(f"Row {row_number}: die/work name already exists ({name}).")
                continue
            try:
                rate = Decimal(str(raw_rate)).quantize(Decimal("0.01"))
                if rate <= 0:
                    raise InvalidOperation
            except (InvalidOperation, ValueError, TypeError):
                errors.append(f"Row {row_number}: rate must be greater than zero.")
                continue
            if code and (code in seen_codes or code in existing_codes):
                errors.append(f"Row {row_number}: die code already exists ({code}).")
                continue

            if not code:
                while f"SL{next_code_number:03d}" in existing_codes or f"SL{next_code_number:03d}" in seen_codes:
                    next_code_number += 1
                code = f"SL{next_code_number:03d}"
                next_code_number += 1

            seen_names.add(name.casefold())
            seen_codes.add(code)
            prepared.append({"name": name, "rate": rate, "is_active": is_active, "die_code": code})

        if errors:
            return Response({"detail": "Import was not completed. Fix the Excel rows and try again.", "errors": errors[:50]}, status=status.HTTP_400_BAD_REQUEST)
        if not prepared:
            return Response({"detail": "No valid die rows were found."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            DiePrice.objects.bulk_create([
                DiePrice(created_by=request.user, **item)
                for item in prepared
            ], batch_size=1000)

        return Response({"detail": f"{len(prepared)} die/work item(s) imported successfully.", "created": len(prepared)}, status=status.HTTP_201_CREATED)


class DiePriceExcelTemplateView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated, IsSuperuser]

    def get(self, request):
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Die Prices"
        headers = ["die_no", "rate"]
        sheet.append(headers)
        sheet.append(["09", 125.00])
        sheet.append(["10", 250.00])

        header_fill = PatternFill("solid", fgColor="344012")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 15
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=2).number_format = "0.00"

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Column", "Required", "Format / example"])
        instructions.append(["die_no", "Yes", "Enter only the number, e.g. 09 or 7000. DIE - is added automatically."])
        instructions.append(["rate", "Yes", "Rate per piece, e.g. 125.00"])
        instructions.append([])
        instructions.append(["Important", "", "SL No. is generated automatically as SL001, SL002, etc."])
        instructions.append(["Upload", "", "Delete the example rows, add your die data, save as .xlsx, and upload it in Master / Die Price."])
        for cell in instructions[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        instructions.column_dimensions["A"].width = 20
        instructions.column_dimensions["B"].width = 12
        instructions.column_dimensions["C"].width = 100

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="die-price-import-template.xlsx"'
        return response
